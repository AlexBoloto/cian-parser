import requests
import os
from bs4 import BeautifulSoup
import re
import numpy as np
import datetime
from fake_useragent import UserAgent
import itertools
import random
import pandas as pd
from time import sleep
from openpyxl import load_workbook
import pickle
maxpage = 80
random.seed(datetime.datetime.now())
name = 0
ids=[6233]
#ids = [16751,19520,7990,5198]  # Ленинский рйаон
#ids = [4501,19520,5094,5198,5640,5902]
#ids = [6180,34597,6233,39537,11712,7484,7182,48687,49046,49135,49182,8160,51271] # Люберцы
#ids = [6233,39537,11712,7484,7182,48687,49046,49135,49182,8160,51271] # Люберцы

def save_cookies(requests_cookiejar, filename):
    with open(filename, 'wb') as f:
        pickle.dump(requests_cookiejar, f)

def load_cookies(filename):
    with open(filename, 'rb') as f:
        return pickle.load(f)

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
def load_proxy():
    url = 'https://www.ip-adress.com/proxy-list'
    r = requests.get(url)
    return r.text
def proxxx(text):
    ip=[]
    tp=[]
    ip_new=[]
    soup = BeautifulSoup(text, 'lxml')
    table = soup.find('table', class_="htable proxylist")
    a = table.find_all('a')
    td = table.find_all('td')
    for i in a:
        ip.append(i.text)
    for i in range(len(td)):
        if (re.findall(r':\d{2,5}',str(td[i].text))):
            tp.append(re.findall(r':\d{2,5}',str(td[i].text)))
        else:
            continue
    port_merged = list(itertools.chain(*tp))
    for i in range(len(ip)):
        ip_new.append(str(ip[i])+str(port_merged[i]))
    random.shuffle(ip_new)
    for proxy in ip_new:
         url = 'http://' + proxy
         try:
             r=requests.get('http://cian.ru',proxies={'http':url})
             if r.status_code == 200:
                 return proxy
         except requests.exceptions.ConnectionError:
             continue
def load_data_agent_and_developer(id,page,proxy):
    url = 'https://www.cian.ru/cat.php?deal_type=sale&engine_version=2&from_developer=2&newobject[0]=%d&offer_type=flat&p=%d'\
        % (id,page)
    cookie = {'session_region_id':'175051','session_main_town_region_id':'175604','cto_lwid':'4df03bb1-c620-43d1-89c8-11f5d850f8e8','_CIAN_GK':'fdbc8a84-4154-4012-80c2-8ffb265494f2',
              'cfids140':'cAnZ8FaIpN4qTEVLNXDvhozQaMSnpGTZ5miWRKrCLHwBWHtmCMv7+welAfBu+BugDF1lw+qn4WG7AIZEwCrQ/kKTcRAlqi1LbJwXUdjNFMb3dVU3IzOV9grQWWjG+KteeNIDbbOCunIZ3g7H9d0svSKQ6xS1XsWyJs04GE8k844='}
    r = requests.get(url,headers={'User-Agent': UserAgent().random},proxies={'http':proxy},cookies=cookie)
    return r.text
def load_data_developer_proxy(id,page,proxy):
    url = \
        'https://www.cian.ru/cat.php?deal_type=sale&engine_version=2&from_developer=1&newobject[0]=%d&offer_type=flat&p=%d'\
        % (id,page)
    #cookie = {'tmr_detect':'1%7C1551366696330','session_region_id':'5953','session_main_town_region_id':'175231','cto_lwid':'15116edd-c5eb-47be-85c5-6613ea1b606d','cto_idcpy':'2e25e6da-20ff-41e8-9574-c11bbe084ab6',
              #'cfids140':'WvFwhg+aeZM1Ll7P7ap8v9q09GU/XE/m9ud67AhzUkomVAceKq19HMzDAf9YRQzApRR/UeTARxO1F68PLpBU8ZALQXE1DUprk8D0zsYFmrL9lIAYGdRN8ZZnkw6CUbQCH1n4O2oc8ml2YmTEHK5IbuFD6KT9mxL0rjVTUAPp1P0='}
    if os.path.exists('cookies.pkl'):
        cookie = load_cookies('cookies.pkl')
    else:
        r = requests.get('https://cian.ru/', headers={'User-Agent': UserAgent().random}, proxies={'http': proxy})
        cookie = r.cookies
        save_cookies(r.cookies,'cookies.pkl')
    r = requests.get(url,headers={'User-Agent': UserAgent().random},proxies={'http':proxy},cookies=cookie)
    return r.text


def load_data_developer(id,page):
    url = \
        'https://www.cian.ru/cat.php?deal_type=sale&engine_version=2&from_developer=1&newobject[0]=%d&offer_type=flat&p=%d'\
        % (id,page)
    cookie = {'tmr_detect':'1%7C1551366696330','session_region_id':'5953','session_main_town_region_id':'175231','cto_lwid':'15116edd-c5eb-47be-85c5-6613ea1b606d','cto_idcpy':'2e25e6da-20ff-41e8-9574-c11bbe084ab6',
              'cfids140':'WvFwhg+aeZM1Ll7P7ap8v9q09GU/XE/m9ud67AhzUkomVAceKq19HMzDAf9YRQzApRR/UeTARxO1F68PLpBU8ZALQXE1DUprk8D0zsYFmrL9lIAYGdRN8ZZnkw6CUbQCH1n4O2oc8ml2YmTEHK5IbuFD6KT9mxL0rjVTUAPp1P0='}
    r = requests.get(url,headers={'User-Agent': UserAgent().random},cookies=cookie)
    return r.text


def cian(text):
    global maxpage
    global name
    soup = BeautifulSoup(text, 'lxml')
    price,page,title,flat_type,flat_sqr,floor,id,data,urls,names,dd = [],[],[],[],[],[],[],[],[],[],[]
    div_price = soup.find_all('div',class_ = re.compile('c6e8ba5398--header'))
    div_max_page =  soup.find_all('a',class_ = re.compile('_93444fe79c-list'))
    div_title = soup.find_all('div',class_ = re.compile('c6e8ba5398--container--F3yyv'))
    div_room_sqr_floor_top = soup.find_all('div',class_ = re.compile('c6e8ba5398--single'))
    div_room_sqr_floor = soup.find_all('div',class_ =re.compile('c6e8ba5398--title--2CW78'))
    links = soup.find_all('a',class_=re.compile('c6e8ba5398--header'))
    for i in links:
        urls.append(i.attrs['href'])
    name = soup.find('div', class_ = re.compile('_93444fe79c-content-title')).a.get_text()
    div_name = soup.find_all('a',class_='c6e8ba5398--building-link--1dQyE')
    div_deadline = soup.find_all('div',class_='c6e8ba5398--deadline--3mUGe')
    zastroyshik = soup.find('a', class_='_93444fe79c-name--1iqIl').get_text()
    for i in div_price:
        price.append(re.search(('[0-9]*[.,]?[0-9]?[0-9]'), re.sub('\s','',i.get_text())).group(0))
    for i in div_name:
        names.append(i.get_text())
    if div_room_sqr_floor_top != [] and div_room_sqr_floor_top is not None:
        for i in div_room_sqr_floor_top:
            a = re.split(',', i.text)
            flat_sqr.append(re.sub('\D','',a[1]))
            flat_type.append(re.sub('\D','',a[0].replace('Студия','0')))
            floor.append(re.sub('\s','',a[2].replace('этаж','')))
            id.append(re.sub('\D', '', i.parent.get('href')))
    if div_room_sqr_floor != []:
        for i in div_room_sqr_floor:
            a = re.split(',', i.text)
            flat_sqr.append(re.sub('\D','',a[1]))
            flat_type.append(re.sub('\D', '', a[0].replace('Студия', '0')))
            floor.append(re.sub('\s','',a[2].replace('этаж','')))
            id.append(re.sub('\D','',i.parent.get('href')))
    else:
        for i in range(len(price)-3):
            flat_sqr.append('')
            flat_type.append('')
            floor.append('')
            id.append('')
    for i in div_max_page:
        if i.text.isdigit():
            page.append(int(i.get_text()))
    try:
        maxpage = max(page)
    except:
        maxpage = 1
    for i in div_title:
        title.append(re.sub('\n\n*','',i.get_text().replace('... Подробнее','')))
    if div_deadline != []:
        for i in div_deadline:
            dd.append(i.get_text().replace('Сдача ГК: ',''))
    else:
        dd = ['' for i in range(len(price))]
    for i in range(len(price)):
        try:
            if names[i] == name:
                try:
                    data.append({
                        'ID':id[i],
                        'Комнат':int(flat_type[i].replace('-комн. апарт.','')),
                        'Площадь':round(float(flat_sqr[i]),2),
                        'Стоимость':int(price[i]),
                        'Цена за метр':round((float(price[i])/float(flat_sqr[i])),2) - otdelka(zastroyshik),
                        'Этаж':floor[i],
                        'ЖК':name,
                        'Срок сдачи':dd[i],
                        'Застройщик':zastroyshik,
                        'Описание':title[i],
                        'Ссылка':urls[i]
                    })
                except:
                    print('Banned')
                    continue
        except:
            print('Levaya dich')
            continue
    print(data)
    return data


def otdelka(zastr):
    df = pd.read_excel('otdelka.xlsx')
    try:
        val = df[df['Застройщик'] == zastr].iloc[0]['Цена']
        return val
    except IndexError:
        return 0

if __name__ == '__main__':
    list_of_zhk = []
    global proxy
    while (True):
        try:
            print('Enter number: 1 - Luber, 2 - Leninskiy, 3 - read from XLSX')
            var = int(input())
            if var == 1:
                list_of_zhk = [6180, 34597, 6233, 39537, 11712, 7484, 7182, 48687, 49046, 49135, 49182, 8160, 51271]  # Luber
                break
            elif var == 2:
                list_of_zhk = [16751, 19520, 7990, 5198]  # Leninskiy
                break
            elif var == 3:
                df = pd.read_excel('id.xlsx')
                for i in range(len(df)):
                    list_of_zhk.append(df.iloc[i]['ID'])
                break
            else:
                print('Incorrect value. Try again')
        except ValueError:
            print('Incorrect value. Try again')
    print('OK')
    while (True):
        try:
            print('Enter proxy')
            proxy = input()
            if ':' in proxy:
                break
            else:
                print('Incorrect value. Try again')
        except ValueError:
            print('Incorrect value. Try again')
    print('OK')
    if os.path.exists('cookies.pkl'):
        while(True):
            try:
                print('Удалить старые cookies? Y/N')
                answer = input()
                if answer.upper() == 'Y':
                    os.remove('cookies.pkl')
                elif answer.upper() == 'N':
                    break
            except:
                print('Неверный ответ')
    for j in range(len(list_of_zhk)):
        data = []
        a = []
        while 1:
            try:
                i=1
                while (i<=maxpage):
                    text = load_data_developer_proxy(list_of_zhk[j],i,proxy)
                    sp = BeautifulSoup(text,'lxml')
                    if sp.title.text == 'Captcha - база объявлений ЦИАН':
                        print('Captcha. Меняем proxy и cookies')
                    data.append(pd.DataFrame(cian(text)))
                    print(i, maxpage)
                    i+=1
                    sleep(random.randint(20,30))
            except:
                continue
            break
        df_res = pd.concat(data,ignore_index=True)
        if maxpage>1:
            df_res.drop_duplicates(subset=['ID'], inplace=True)
        print(df_res)
        for i in range(5):
            try:
                df_1 = df_res[df_res['Комнат'] == i]
                a.append({
                    'Название ЖК': name,
                    'Комнат':i,
                    'Максимальная площадь': df_1['Площадь'].max(axis=0),
                    'Средняя площадь': (df_1['Площадь'].mean(axis=0)).round(2),
                    'Минимальная площадь': df_1['Площадь'].min(axis=0),
                    'Максимальная стоимость': df_1['Стоимость'].max(axis=0),
                    'Средняя стоимость': (df_1['Стоимость'].mean(axis=0)).round(2),
                    'Минимальная стоимость': df_1['Стоимость'].min(axis=0),
                    'Максимальная цена за метр': (df_1['Цена за метр'].max(axis=0)).round(2),
                    'Средняя взвешенная цена за метр': (np.average(df_1['Цена за метр'], weights=df_1['Площадь'])).round(2),
                    'Минимальная цена за метр': (df_1['Цена за метр'].min(axis=0)).round(2),
                    'Объем выборки': df_1['Стоимость'].count()
                })
            except AttributeError:
                print('Attribute Error')
                continue
        analytics = pd.DataFrame(a)
        analytics = analytics[['Название ЖК','Комнат','Минимальная площадь', 'Средняя площадь', 'Максимальная площадь', 'Минимальная цена за метр',
                   'Средняя взвешенная цена за метр', 'Максимальная цена за метр',
                   'Минимальная стоимость', 'Средняя стоимость', 'Максимальная стоимость', 'Объем выборки']]
        print(analytics)
        if os.path.exists('C:\\Python\\untitled33333\\Аналитика на ' + datetime.date.today().strftime("%d-%m-%Y") + '.xlsx'):
            append_df_to_excel('Аналитика Leninskiy на ' + datetime.date.today().strftime("%d-%m-%Y") + '.xlsx',df_res,header=False,sheet_name='Проекты',index=False)
            append_df_to_excel('Аналитика Leninskiy на ' + datetime.date.today().strftime("%d-%m-%Y") + '.xlsx', analytics,
                               header=False, sheet_name='Аналитика', index=False)
        else:
            append_df_to_excel('Аналитика Leninskiy на ' + datetime.date.today().strftime("%d-%m-%Y") + '.xlsx',df_res,sheet_name='Проекты',index=False)
            append_df_to_excel('Аналитика Leninskiy на ' + datetime.date.today().strftime("%d-%m-%Y") + '.xlsx',analytics,sheet_name='Аналитика',index=False)